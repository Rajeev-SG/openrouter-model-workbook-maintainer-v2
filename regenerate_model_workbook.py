#!/usr/bin/env python3
"""
Regenerate an OpenRouter / Artificial Analysis / Vals AI model comparison workbook.

What this script does
---------------------
1. Fetches current model/pricing metadata from OpenRouter.
2. Fetches current model-level benchmark/speed/pricing metadata from Artificial Analysis.
3. Scrapes selected Artificial Analysis provider pages for provider-specific stats
   such as fastest provider, lowest latency provider, and cheapest provider.
4. Scrapes selected Vals AI model pages for Vals Index / model-guide metrics and
   per-benchmark ranks.
5. Builds an XLSX workbook with the same sheet structure as the workbook previously
   produced in this conversation:
      - Overview
      - Pricing_OpenRouter
      - ArtificialAnalysis
      - ValsAI
      - Vals_Benchmarks
      - Sources_Notes

Why the script mixes APIs and scraping
--------------------------------------
- OpenRouter provides a public models API, which is the best source of truth for
  slug, context length, and token pricing.
- Artificial Analysis provides a free API for model-level data. Provider-breakout
  stats are easiest to extract from the public provider pages.
- Vals AI exposes the metrics needed on public model pages; this script scrapes those
  pages because a comparable public API is not documented in the sources used here.

Determinism / reproducibility
-----------------------------
- All fetched API responses and page HTML are cached to disk under --cache-dir.
- Re-running with the same cache files and without --refresh gives the same source
  inputs, so the workbook should be reproducible.
- The crosswalk between OpenRouter, Artificial Analysis, and Vals is encoded in
  DEFAULT_MODEL_MAP or an optional mapping CSV file.

Requirements
------------
pip install requests beautifulsoup4 openpyxl python-dateutil

Environment variables
---------------------
AA_API_KEY=...                 required for Artificial Analysis API
OPENROUTER_API_KEY=...         optional for OpenRouter models API if required

Usage examples
--------------
python regenerate_model_workbook.py
python regenerate_model_workbook.py --out updated.xlsx --refresh
python regenerate_model_workbook.py --mapping-csv model_map.csv --out custom.xlsx
"""
from __future__ import annotations

import argparse
import csv
import json
import os
import re
import sys
import time
from dataclasses import dataclass, field
from datetime import datetime, timezone
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Tuple

import requests
from bs4 import BeautifulSoup
from dateutil import parser as dtparser
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


# ---------------------------
# Config / model crosswalk
# ---------------------------

@dataclass
class AAVariant:
    variant: str
    model_slug: str
    creator_slug: Optional[str] = None
    provider_url: Optional[str] = None
    intelligence_url: Optional[str] = None
    preferred: bool = False


@dataclass
class ValsVariant:
    variant: str
    model_url: str
    preferred: bool = False


@dataclass
class ModelSpec:
    family: str
    openrouter_slug: str
    openrouter_page_url: str
    notes: str = ""
    aa_variants: List[AAVariant] = field(default_factory=list)
    vals_variants: List[ValsVariant] = field(default_factory=list)


DEFAULT_MODEL_MAP: List[ModelSpec] = [
    ModelSpec(
        family="Moonshot AI Kimi K2.5",
        openrouter_slug="moonshotai/kimi-k2.5",
        openrouter_page_url="https://openrouter.ai/moonshotai/kimi-k2.5/pricing",
        notes="Kimi K2.5 family.",
        aa_variants=[
            AAVariant(
                variant="Reasoning / Thinking",
                model_slug="kimi-k2-5",
                creator_slug="kimi",
                provider_url="https://artificialanalysis.ai/models/kimi-k2-5/providers",
                intelligence_url="https://artificialanalysis.ai/models/kimi-k2-5",
                preferred=True,
            ),
            AAVariant(
                variant="Non-reasoning",
                model_slug="kimi-k2-5-non-reasoning",
                creator_slug="kimi",
                provider_url="https://artificialanalysis.ai/models/kimi-k2-5-non-reasoning/providers",
                intelligence_url="https://artificialanalysis.ai/models/kimi-k2-5-non-reasoning",
                preferred=False,
            ),
        ],
        vals_variants=[
            ValsVariant(
                variant="Thinking",
                model_url="https://www.vals.ai/models/kimi_kimi-k2.5-thinking",
                preferred=True,
            )
        ],
    ),
    ModelSpec(
        family="DeepSeek V3.2",
        openrouter_slug="deepseek/deepseek-v3.2",
        openrouter_page_url="https://openrouter.ai/deepseek/deepseek-v3.2",
        notes="DeepSeek V3.2 family.",
        aa_variants=[
            AAVariant(
                variant="Reasoning / Thinking",
                model_slug="deepseek-v3-2-reasoning",
                creator_slug="deepseek",
                provider_url="https://artificialanalysis.ai/models/deepseek-v3-2-reasoning/providers",
                intelligence_url="https://artificialanalysis.ai/models/deepseek-v3-2-reasoning",
                preferred=True,
            ),
            AAVariant(
                variant="Non-reasoning",
                model_slug="deepseek-v3-2",
                creator_slug="deepseek",
                provider_url="https://artificialanalysis.ai/models/deepseek-v3-2/providers",
                intelligence_url="https://artificialanalysis.ai/models/deepseek-v3-2",
                preferred=False,
            ),
        ],
        vals_variants=[
            ValsVariant(
                variant="Nonthinking",
                model_url="https://www.vals.ai/models/fireworks_deepseek-v3p2",
                preferred=False,
            ),
            ValsVariant(
                variant="Thinking",
                model_url="https://www.vals.ai/models/fireworks_deepseek-v3p2-thinking",
                preferred=True,
            ),
        ],
    ),
    ModelSpec(
        family="Google Gemini 3 Flash Preview",
        openrouter_slug="google/gemini-3-flash-preview",
        openrouter_page_url="https://openrouter.ai/google/gemini-3-flash-preview",
        notes="Gemini 3 Flash Preview family.",
        aa_variants=[
            AAVariant(
                variant="Reasoning",
                model_slug="gemini-3-flash-reasoning",
                creator_slug="google",
                provider_url="https://artificialanalysis.ai/models/gemini-3-flash-reasoning/providers",
                intelligence_url="https://artificialanalysis.ai/models/gemini-3-flash-reasoning",
                preferred=True,
            ),
            AAVariant(
                variant="Non-reasoning",
                model_slug="gemini-3-flash",
                creator_slug="google",
                provider_url="https://artificialanalysis.ai/models/gemini-3-flash/providers",
                intelligence_url="https://artificialanalysis.ai/models/gemini-3-flash",
                preferred=False,
            ),
        ],
        vals_variants=[
            ValsVariant(
                variant="Preview",
                model_url="https://www.vals.ai/models/google_gemini-3-flash-preview",
                preferred=True,
            )
        ],
    ),
    ModelSpec(
        family="OpenAI GPT-5 Mini",
        openrouter_slug="openai/gpt-5-mini",
        openrouter_page_url="https://openrouter.ai/openai/gpt-5-mini",
        notes="GPT-5 Mini family.",
        aa_variants=[
            AAVariant(
                variant="High reasoning",
                model_slug="gpt-5-mini",
                creator_slug="openai",
                provider_url="https://artificialanalysis.ai/models/gpt-5-mini/providers",
                intelligence_url="https://artificialanalysis.ai/models/gpt-5-mini",
                preferred=True,
            ),
        ],
        vals_variants=[
            ValsVariant(
                variant="Default",
                model_url="https://www.vals.ai/models/openai_gpt-5-mini-2025-08-07",
                preferred=True,
            )
        ],
    ),
    ModelSpec(
        family="MiniMax M2.7",
        openrouter_slug="minimax/minimax-m2.7",
        openrouter_page_url="https://openrouter.ai/minimax/minimax-m2.7",
        notes="MiniMax M2.7 family.",
        aa_variants=[
            AAVariant(
                variant="Standard",
                model_slug="minimax-m2-7",
                creator_slug="minimax",
                provider_url="https://artificialanalysis.ai/models/minimax-m2-7/providers",
                intelligence_url="https://artificialanalysis.ai/models/minimax-m2-7",
                preferred=True,
            ),
        ],
        vals_variants=[
            ValsVariant(
                variant="Default",
                model_url="https://www.vals.ai/models/minimax_MiniMax-M2.7",
                preferred=True,
            )
        ],
    ),
]


# ---------------------------
# Utility helpers
# ---------------------------

def slugify_filename(value: str) -> str:
    return re.sub(r"[^a-zA-Z0-9._-]+", "_", value).strip("_")


def ensure_dir(path: Path) -> None:
    path.mkdir(parents=True, exist_ok=True)


def parse_decimal(value: Any) -> Optional[float]:
    if value is None or value == "":
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip()
    text = text.replace(",", "")
    text = text.replace("$", "")
    try:
        return float(Decimal(text))
    except (InvalidOperation, ValueError):
        return None


def parse_k_number(text: str) -> Optional[int]:
    """
    Convert values like '262k', '1M', '1,048,576', '128000' to integer tokens.
    """
    if not text:
        return None
    raw = text.strip().replace(",", "").lower()
    try:
        if raw.endswith("k"):
            return int(float(raw[:-1]) * 1000)
        if raw.endswith("m"):
            return int(float(raw[:-1]) * 1_000_000)
        return int(float(raw))
    except ValueError:
        return None


def parse_date_loose(text: Optional[str]) -> Optional[str]:
    if not text:
        return None
    try:
        dt = dtparser.parse(text, fuzzy=True)
        return dt.date().isoformat()
    except Exception:
        return None


def first_or_none(values: Iterable[Any]) -> Any:
    for v in values:
        if v not in (None, "", []):
            return v
    return None


def format_float(value: Optional[float], digits: int = 2) -> Optional[float]:
    if value is None:
        return None
    return round(float(value), digits)


def now_utc_iso() -> str:
    return datetime.now(timezone.utc).replace(microsecond=0).isoformat()


# ---------------------------
# Fetch / cache layer
# ---------------------------

class Fetcher:
    def __init__(self, cache_dir: Path, refresh: bool = False, timeout: int = 60):
        self.cache_dir = cache_dir
        self.refresh = refresh
        self.timeout = timeout
        self.session = requests.Session()
        self.session.headers.update(
            {
                "User-Agent": "model-workbook-regenerator/1.0 (+deterministic spreadsheet build)"
            }
        )

    def _cache_path(self, category: str, name: str, ext: str) -> Path:
        path = self.cache_dir / category / f"{slugify_filename(name)}.{ext}"
        ensure_dir(path.parent)
        return path

    def get_text(self, url: str, category: str, name: Optional[str] = None, headers: Optional[Dict[str, str]] = None) -> str:
        cache_path = self._cache_path(category, name or url, "html")
        if cache_path.exists() and not self.refresh:
            return cache_path.read_text(encoding="utf-8")
        response = self.session.get(url, headers=headers or {}, timeout=self.timeout)
        response.raise_for_status()
        text = response.text
        cache_path.write_text(text, encoding="utf-8")
        return text

    def get_json(self, url: str, category: str, name: Optional[str] = None, headers: Optional[Dict[str, str]] = None) -> Dict[str, Any]:
        cache_path = self._cache_path(category, name or url, "json")
        if cache_path.exists() and not self.refresh:
            return json.loads(cache_path.read_text(encoding="utf-8"))
        response = self.session.get(url, headers=headers or {}, timeout=self.timeout)
        response.raise_for_status()
        payload = response.json()
        cache_path.write_text(json.dumps(payload, indent=2, sort_keys=True), encoding="utf-8")
        return payload


# ---------------------------
# Parsing helpers
# ---------------------------

def html_to_text(html: str) -> str:
    soup = BeautifulSoup(html, "html.parser")
    text = soup.get_text("\n", strip=True)
    text = re.sub(r"\n{2,}", "\n", text)
    return text


def parse_openrouter_models_api(payload: Dict[str, Any]) -> Dict[str, Dict[str, Any]]:
    rows: Dict[str, Dict[str, Any]] = {}
    for item in payload.get("data", []):
        slug = item.get("canonical_slug") or item.get("id")
        if not slug:
            continue
        pricing = item.get("pricing") or {}
        rows[slug] = {
            "slug": slug,
            "name": item.get("name"),
            "created_epoch": item.get("created"),
            "context_length": item.get("context_length"),
            "prompt_price_per_token": parse_decimal(pricing.get("prompt")),
            "completion_price_per_token": parse_decimal(pricing.get("completion")),
            "request_price": parse_decimal(pricing.get("request")),
            "image_price": parse_decimal(pricing.get("image")),
            "architecture": item.get("architecture"),
            "supported_parameters": item.get("supported_parameters"),
        }
    return rows


def parse_openrouter_page(html: str) -> Dict[str, Any]:
    text = html_to_text(html)
    data: Dict[str, Any] = {"release_date": None, "context_tokens": None, "input_m": None, "output_m": None, "audio_input_m": None, "web_search_k": None}

    m = re.search(r"Released\s+([A-Za-z]{3,9}\s+\d{1,2},\s+\d{4})", text)
    if m:
        data["release_date"] = parse_date_loose(m.group(1))

    m = re.search(r"Released\s+[A-Za-z]{3,9}\s+\d{1,2},\s+\d{4}\s+([\d,]+)\s+context", text)
    if m:
        data["context_tokens"] = parse_k_number(m.group(1))

    m = re.search(r"\$([\d.]+)\/M\s+input tokens", text, re.IGNORECASE)
    if m:
        data["input_m"] = parse_decimal(m.group(1))

    m = re.search(r"\$([\d.]+)\/M\s+output tokens", text, re.IGNORECASE)
    if m:
        data["output_m"] = parse_decimal(m.group(1))

    m = re.search(r"\$([\d.]+)\/M\s+audio input", text, re.IGNORECASE)
    if m:
        data["audio_input_m"] = parse_decimal(m.group(1))

    m = re.search(r"\$([\d.]+)\/K\s+web search", text, re.IGNORECASE)
    if m:
        data["web_search_k"] = parse_decimal(m.group(1))

    return data


def parse_aa_models_api(payload: Dict[str, Any]) -> Dict[Tuple[Optional[str], str], Dict[str, Any]]:
    """
    Keyed by (creator_slug, model_slug) where creator_slug may be None if absent.
    """
    rows: Dict[Tuple[Optional[str], str], Dict[str, Any]] = {}
    for item in payload.get("data", []):
        creator = item.get("model_creator") or {}
        creator_slug = creator.get("slug")
        slug = item.get("slug")
        if not slug:
            continue
        evaluations = item.get("evaluations") or {}
        pricing = item.get("pricing") or {}
        rows[(creator_slug, slug)] = {
            "id": item.get("id"),
            "name": item.get("name"),
            "slug": slug,
            "creator_slug": creator_slug,
            "creator_name": creator.get("name"),
            "intelligence_index": parse_decimal(evaluations.get("artificial_analysis_intelligence_index")),
            "coding_index": parse_decimal(evaluations.get("artificial_analysis_coding_index")),
            "math_index": parse_decimal(evaluations.get("artificial_analysis_math_index")),
            "aa_blended_price": parse_decimal(pricing.get("price_1m_blended_3_to_1")),
            "aa_input_price": parse_decimal(pricing.get("price_1m_input_tokens")),
            "aa_output_price": parse_decimal(pricing.get("price_1m_output_tokens")),
            "median_tps": parse_decimal(item.get("median_output_tokens_per_second")),
            "median_ttft_seconds": parse_decimal(item.get("median_time_to_first_token_seconds")),
        }
    return rows


def parse_aa_provider_page(html: str) -> Dict[str, Any]:
    """
    Parse provider summary stats from an Artificial Analysis provider page.

    The page text typically contains sentences such as:
    - "The most affordable providers for Kimi K2.5 (Reasoning) by blended price are DeepInfra ($0.90 per 1M tokens), ..."
    - "Eigen AI is the fastest at 471.1 t/s"
    - "The provider with the lowest latency is Baseten at 0.37 seconds"
    - JSON / Function calling provider counts may appear in text or tables.
    """
    text = html_to_text(html)
    normalized = re.sub(r"\s+", " ", text)

    result: Dict[str, Any] = {
        "fastest_provider": None,
        "fastest_tps": None,
        "lowest_latency_provider": None,
        "lowest_latency_seconds": None,
        "cheapest_provider": None,
        "cheapest_blended_price": None,
        "lowest_input_price": None,
        "lowest_output_price": None,
        "json_support": None,
        "function_calling": None,
    }

    m = re.search(r"most affordable provider\s*s? for .*? are ([^($]+)\(\$([\d.]+) per 1M tokens\)", normalized, re.IGNORECASE)
    if m:
        result["cheapest_provider"] = m.group(1).strip().rstrip(",")
        result["cheapest_blended_price"] = parse_decimal(m.group(2))
    else:
        m = re.search(r"most affordable is ([^$]+?) at \$([\d.]+) per 1M tokens", normalized, re.IGNORECASE)
        if m:
            result["cheapest_provider"] = m.group(1).strip().rstrip(",")
            result["cheapest_blended_price"] = parse_decimal(m.group(2))

    m = re.search(r"lowest input token pricing .*? are [^$]*\(\$([\d.]+) per 1M input tokens\)", normalized, re.IGNORECASE)
    if m:
        result["lowest_input_price"] = parse_decimal(m.group(1))
    m = re.search(r"lowest output token pricing .*? are [^$]*\(\$([\d.]+) per 1M output tokens\)", normalized, re.IGNORECASE)
    if m:
        result["lowest_output_price"] = parse_decimal(m.group(1))

    m = re.search(r"([A-Za-z0-9 .&/+_-]+?) is the fastest at ([\d.]+)\s*t/s", normalized, re.IGNORECASE)
    if m:
        result["fastest_provider"] = m.group(1).strip().rstrip(",")
        result["fastest_tps"] = parse_decimal(m.group(2))

    m = re.search(r"(?:lowest latency is|fastest time to first token is|provider with the lowest latency is)\s*([A-Za-z0-9 .&/+_-]+?)\s*at\s*([\d.]+)\s*seconds", normalized, re.IGNORECASE)
    if m:
        result["lowest_latency_provider"] = m.group(1).strip().rstrip(",")
        result["lowest_latency_seconds"] = parse_decimal(m.group(2))

    # Counts like "15 / 17 providers"
    json_match = re.search(r"JSON(?: output| mode)?[^0-9]*(\d+\s*/\s*\d+\s*providers?)", normalized, re.IGNORECASE)
    if json_match:
        result["json_support"] = json_match.group(1).replace(" providers", "").replace(" provider", "")
    func_match = re.search(r"Function Calling[^0-9]*(\d+\s*/\s*\d+\s*providers?)", normalized, re.IGNORECASE)
    if func_match:
        result["function_calling"] = func_match.group(1).replace(" providers", "").replace(" provider", "")

    # Fallback: tables may present labels and values without prose.
    # Search for standalone labels if earlier regexes fail.
    if result["json_support"] is None:
        m = re.search(r"JSON(?: output| mode)?\s+(\d+\s*/\s*\d+)", normalized, re.IGNORECASE)
        if m:
            result["json_support"] = m.group(1)
    if result["function_calling"] is None:
        m = re.search(r"Function Calling\s+(\d+\s*/\s*\d+)", normalized, re.IGNORECASE)
        if m:
            result["function_calling"] = m.group(1)

    return result


def parse_vals_page(html: str) -> Dict[str, Any]:
    """
    Parse a Vals model page. The public pages expose model-guide metrics such as:
    - Release Date
    - Accuracy (Vals Index)
    - Latency (Vals Index)
    - Cost/Test (Vals Index) or Avg. Cost (In/Out)
    - Context Window
    - Max Output Tokens
    - Default Provider
    - Reasoning Effort
    - benchmark ranks e.g. "SWE-bench 10/62"
    """
    text = html_to_text(html)
    normalized = re.sub(r"\s+", " ", text)

    data: Dict[str, Any] = {
        "release_date": None,
        "accuracy": None,
        "ci_plus_minus": None,
        "latency_seconds": None,
        "cost_per_test": None,
        "avg_cost_input": None,
        "avg_cost_output": None,
        "context_tokens": None,
        "max_output_tokens": None,
        "default_provider": None,
        "reasoning_effort": None,
        "vals_index_rank": None,
        "vals_index_population": None,
        "benchmark_highlights": None,
        "benchmarks": [],
    }

    m = re.search(r"Release Date:\s*([0-9]{1,2}/[0-9]{1,2}/[0-9]{4})", normalized, re.IGNORECASE)
    if m:
        data["release_date"] = parse_date_loose(m.group(1))

    m = re.search(r"Accuracy\s*\(Vals Index\)\s*([\d.]+)%\s*±\s*([\d.]+)", normalized, re.IGNORECASE)
    if m:
        data["accuracy"] = parse_decimal(m.group(1)) / 100.0
        data["ci_plus_minus"] = parse_decimal(m.group(2)) / 100.0
    else:
        m = re.search(r"Accuracy\s*\(Average\)\s*([\d.]+)%", normalized, re.IGNORECASE)
        if m:
            data["accuracy"] = parse_decimal(m.group(1)) / 100.0

    m = re.search(r"Latency\s*\((?:Vals Index|Average)\)\s*([\d.]+)s", normalized, re.IGNORECASE)
    if m:
        data["latency_seconds"] = parse_decimal(m.group(1))

    m = re.search(r"Cost/Test\s*\(Vals Index\)\s*\$([\d.]+)", normalized, re.IGNORECASE)
    if m:
        data["cost_per_test"] = parse_decimal(m.group(1))

    m = re.search(r"Avg\.?\s*Cost\s*\(In/Out\)\s*\$([\d.]+)\s*/\s*\$([\d.]+)", normalized, re.IGNORECASE)
    if m:
        data["avg_cost_input"] = parse_decimal(m.group(1))
        data["avg_cost_output"] = parse_decimal(m.group(2))

    m = re.search(r"Context Window\s*([\d.,]+k?|[\d.]+M)", normalized, re.IGNORECASE)
    if m:
        data["context_tokens"] = parse_k_number(m.group(1))

    m = re.search(r"Max Output Tokens\s*([\d.,]+k?|[\d.]+M)", normalized, re.IGNORECASE)
    if m:
        data["max_output_tokens"] = parse_k_number(m.group(1))

    m = re.search(r"Default Provider\s*:?\s*([A-Za-z0-9 .&/+_-]+)", normalized, re.IGNORECASE)
    if m:
        provider = m.group(1).strip()
        provider = re.split(r"\s{2,}|Temperature|Top P|Top K|Some benchmarks may use", provider)[0].strip()
        data["default_provider"] = provider

    m = re.search(r"Reasoning Effort\s*:?\s*([A-Za-z0-9 .&/+_-]+)", normalized, re.IGNORECASE)
    if m:
        data["reasoning_effort"] = m.group(1).strip()

    # Benchmark-style ranks in prose. Keep all hits and later pick a few highlights.
    # Allow names like "SWE-bench", "Terminal-Bench 2", "Vals Index", "MMLU Pro"
    benchmark_hits: List[Tuple[str, int, int]] = []
    for bench, rank, pop in re.findall(r"([A-Za-z0-9.+/_ -]{2,}?)\s+(\d+)\s*/\s*(\d+)", normalized):
        bench = bench.strip(" -")
        # Filter obvious false positives
        if len(bench) < 3:
            continue
        if bench.lower().startswith(("temperature", "top p", "top k")):
            continue
        rank_i = int(rank)
        pop_i = int(pop)
        if rank_i <= 0 or pop_i <= 0 or rank_i > pop_i:
            continue
        benchmark_hits.append((bench, rank_i, pop_i))

    # Deduplicate while preserving order.
    seen = set()
    cleaned_hits: List[Tuple[str, int, int]] = []
    for item in benchmark_hits:
        key = (item[0].lower(), item[1], item[2])
        if key in seen:
            continue
        seen.add(key)
        cleaned_hits.append(item)
    data["benchmarks"] = cleaned_hits

    # Determine Vals Index rank from either explicit benchmark hit or a prose statement like "#12 (60.14%) on the Vals Index"
    for bench, rank_i, pop_i in cleaned_hits:
        if bench.lower() == "vals index":
            data["vals_index_rank"] = rank_i
            data["vals_index_population"] = pop_i
            break

    if data["vals_index_rank"] is None:
        m = re.search(r"#(\d+)\s*\((?:[\d.]+)%\)\s*on the Vals Index", normalized, re.IGNORECASE)
        if m:
            data["vals_index_rank"] = int(m.group(1))
    if data["vals_index_population"] is None:
        m = re.search(r"Vals Index[^0-9]+#\d+/?(\d+)", normalized, re.IGNORECASE)
        if m:
            data["vals_index_population"] = int(m.group(1))

    # Build benchmark highlights from first 3-4 non-vals-index benchmarks
    highlights = []
    for bench, rank_i, pop_i in cleaned_hits:
        if bench.lower() == "vals index":
            continue
        highlights.append(f"{bench} {rank_i}/{pop_i}")
        if len(highlights) == 4:
            break
    if highlights:
        data["benchmark_highlights"] = "; ".join(highlights)

    return data


# ---------------------------
# Mapping helpers
# ---------------------------

def load_mapping_csv(mapping_csv: Optional[Path]) -> List[ModelSpec]:
    if not mapping_csv:
        return DEFAULT_MODEL_MAP

    # Expected columns:
    # family,openrouter_slug,openrouter_page_url,notes,aa_variant,aa_model_slug,aa_creator_slug,aa_provider_url,aa_intelligence_url,aa_preferred,vals_variant,vals_model_url,vals_preferred
    rows = list(csv.DictReader(mapping_csv.read_text(encoding="utf-8").splitlines()))
    grouped: Dict[Tuple[str, str], ModelSpec] = {}

    for row in rows:
        key = (row["family"], row["openrouter_slug"])
        spec = grouped.get(key)
        if spec is None:
            spec = ModelSpec(
                family=row["family"],
                openrouter_slug=row["openrouter_slug"],
                openrouter_page_url=row["openrouter_page_url"],
                notes=row.get("notes", ""),
            )
            grouped[key] = spec

        aa_variant = row.get("aa_variant")
        if aa_variant:
            spec.aa_variants.append(
                AAVariant(
                    variant=aa_variant,
                    model_slug=row.get("aa_model_slug", ""),
                    creator_slug=row.get("aa_creator_slug") or None,
                    provider_url=row.get("aa_provider_url") or None,
                    intelligence_url=row.get("aa_intelligence_url") or None,
                    preferred=str(row.get("aa_preferred", "")).strip().lower() in ("1", "true", "yes", "y"),
                )
            )

        vals_variant = row.get("vals_variant")
        if vals_variant:
            spec.vals_variants.append(
                ValsVariant(
                    variant=vals_variant,
                    model_url=row.get("vals_model_url", ""),
                    preferred=str(row.get("vals_preferred", "")).strip().lower() in ("1", "true", "yes", "y"),
                )
            )
    return list(grouped.values())


# ---------------------------
# Data collection
# ---------------------------

def collect_openrouter_data(fetcher: Fetcher, specs: List[ModelSpec]) -> Dict[str, Dict[str, Any]]:
    headers = {}
    token = os.getenv("OPENROUTER_API_KEY")
    if token:
        headers["Authorization"] = f"Bearer {token}"

    api_payload = fetcher.get_json(
        "https://openrouter.ai/api/v1/models",
        category="openrouter_api",
        name="models",
        headers=headers if headers else None,
    )
    api_by_slug = parse_openrouter_models_api(api_payload)

    pricing_rows: Dict[str, Dict[str, Any]] = {}
    for spec in specs:
        api_item = api_by_slug.get(spec.openrouter_slug, {})
        page_html = fetcher.get_text(
            spec.openrouter_page_url,
            category="openrouter_pages",
            name=spec.openrouter_slug,
        )
        page_data = parse_openrouter_page(page_html)

        prompt_price_per_token = api_item.get("prompt_price_per_token")
        completion_price_per_token = api_item.get("completion_price_per_token")
        input_m = page_data["input_m"] if page_data["input_m"] is not None else (prompt_price_per_token * 1_000_000 if prompt_price_per_token is not None else None)
        output_m = page_data["output_m"] if page_data["output_m"] is not None else (completion_price_per_token * 1_000_000 if completion_price_per_token is not None else None)

        pricing_rows[spec.family] = {
            "Model Family": spec.family,
            "OpenRouter Slug": spec.openrouter_slug,
            "Release Date": first_or_none([page_data["release_date"], None]),
            "Context Tokens": first_or_none([page_data["context_tokens"], api_item.get("context_length")]),
            "Input $/M Tokens": format_float(input_m, 4) if input_m is not None else None,
            "Output $/M Tokens": format_float(output_m, 4) if output_m is not None else None,
            "Audio Input $/M": format_float(page_data["audio_input_m"], 4) if page_data["audio_input_m"] is not None else None,
            "Web Search $/K": format_float(page_data["web_search_k"], 4) if page_data["web_search_k"] is not None else None,
            "Source URL": spec.openrouter_page_url,
            "Notes": spec.notes,
        }
    return pricing_rows


def collect_aa_data(fetcher: Fetcher, specs: List[ModelSpec]) -> List[Dict[str, Any]]:
    api_key = os.getenv("AA_API_KEY")
    if not api_key:
        raise RuntimeError("AA_API_KEY is required to query the Artificial Analysis API.")

    payload = fetcher.get_json(
        "https://artificialanalysis.ai/api/v2/data/llms/models",
        category="artificialanalysis_api",
        name="llm_models",
        headers={"x-api-key": api_key},
    )
    aa_api = parse_aa_models_api(payload)

    rows: List[Dict[str, Any]] = []
    for spec in specs:
        for variant in spec.aa_variants:
            api_row = aa_api.get((variant.creator_slug, variant.model_slug))
            if api_row is None:
                # fallback: ignore creator
                api_row = next(
                    (v for (creator_slug, slug), v in aa_api.items() if slug == variant.model_slug),
                    None,
                )
            if api_row is None:
                creator_matches = sorted({slug for (creator_slug, slug), _ in aa_api.items() if creator_slug == variant.creator_slug})
                nearby = [slug for slug in creator_matches if variant.model_slug.split("-")[0] in slug][:12]
                suggestions = nearby or creator_matches[:12]
                raise KeyError(
                    f"Could not find Artificial Analysis model slug '{variant.model_slug}' "
                    f"(creator '{variant.creator_slug}') in API response. "
                    f"Update your mapping. Example available slugs for this creator: {', '.join(suggestions) if suggestions else 'none found'}."
                )

            provider_data = {}
            if variant.provider_url:
                provider_html = fetcher.get_text(
                    variant.provider_url,
                    category="artificialanalysis_provider_pages",
                    name=f"{spec.family}_{variant.variant}",
                )
                provider_data = parse_aa_provider_page(provider_html)

            rows.append(
                {
                    "Model Family": spec.family,
                    "Variant": variant.variant,
                    "Intelligence Index": api_row.get("intelligence_index"),
                    "Fastest Provider": provider_data.get("fastest_provider"),
                    "Fastest Tokens/sec": provider_data.get("fastest_tps"),
                    "Lowest Latency Provider": provider_data.get("lowest_latency_provider"),
                    "Lowest Latency (s)": provider_data.get("lowest_latency_seconds"),
                    "Cheapest Provider": provider_data.get("cheapest_provider"),
                    "Cheapest Blended $/M": provider_data.get("cheapest_blended_price"),
                    "Lowest Input $/M": provider_data.get("lowest_input_price") or api_row.get("aa_input_price"),
                    "Lowest Output $/M": provider_data.get("lowest_output_price") or api_row.get("aa_output_price"),
                    "JSON Support": provider_data.get("json_support"),
                    "Function Calling": provider_data.get("function_calling"),
                    "Provider Page URL": variant.provider_url,
                    "Intelligence Source URL": variant.intelligence_url,
                    "Notes": api_row.get("id"),
                    "_preferred": variant.preferred,
                }
            )
    return rows


def collect_vals_data(fetcher: Fetcher, specs: List[ModelSpec]) -> Tuple[List[Dict[str, Any]], List[Dict[str, Any]]]:
    vals_rows: List[Dict[str, Any]] = []
    benchmark_rows: List[Dict[str, Any]] = []

    for spec in specs:
        for variant in spec.vals_variants:
            html = fetcher.get_text(
                variant.model_url,
                category="vals_model_pages",
                name=f"{spec.family}_{variant.variant}",
            )
            parsed = parse_vals_page(html)
            vals_rows.append(
                {
                    "Model Family": spec.family,
                    "Variant": variant.variant,
                    "Release Date": parsed.get("release_date"),
                    "Accuracy %": parsed.get("accuracy"),
                    "CI +/-": parsed.get("ci_plus_minus"),
                    "Latency (s)": parsed.get("latency_seconds"),
                    "Cost/Test $": parsed.get("cost_per_test"),
                    "Avg Cost Input $": parsed.get("avg_cost_input"),
                    "Avg Cost Output $": parsed.get("avg_cost_output"),
                    "Context Tokens": parsed.get("context_tokens"),
                    "Max Output Tokens": parsed.get("max_output_tokens"),
                    "Default Provider": parsed.get("default_provider"),
                    "Reasoning Effort": parsed.get("reasoning_effort"),
                    "Vals Index Rank": parsed.get("vals_index_rank"),
                    "Vals Index Population": parsed.get("vals_index_population"),
                    "Benchmark Highlights": parsed.get("benchmark_highlights"),
                    "Source URL": variant.model_url,
                    "Benchmark Note": "",
                    "_preferred": variant.preferred,
                }
            )
            for bench, rank_i, pop_i in parsed.get("benchmarks", []):
                benchmark_rows.append(
                    {
                        "Model Family": spec.family,
                        "Variant": variant.variant,
                        "Benchmark": bench,
                        "Rank": rank_i,
                        "Population": pop_i,
                        "Source URL": variant.model_url,
                        "Notes": "",
                    }
                )
    return vals_rows, benchmark_rows


# ---------------------------
# Workbook building
# ---------------------------

HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
SECTION_FILL = PatternFill("solid", fgColor="D9EAF7")
THIN_GRAY = Side(style="thin", color="D9D9D9")
BORDER_BOTTOM = Border(bottom=THIN_GRAY)
WHITE_FONT = Font(color="FFFFFF", bold=True)
BOLD_FONT = Font(bold=True)
LINK_FONT = Font(color="008000")
INPUT_FONT = Font(color="0000FF")
FORMULA_FONT = Font(color="000000")
STATIC_FONT = Font(color="666666")


def write_sheet(ws, headers: List[str], rows: List[Dict[str, Any]], formula_columns: Optional[Dict[str, str]] = None) -> None:
    ws.append(headers)
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = WHITE_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER_BOTTOM

    formula_columns = formula_columns or {}
    for excel_row_index, row in enumerate(rows, start=2):
        values = []
        for header in headers:
            if header in formula_columns:
                formula = formula_columns[header].format(row=excel_row_index)
                values.append(formula)
            else:
                values.append(row.get(header))
        ws.append(values)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


def autosize_columns(ws, max_width: int = 45) -> None:
    for col_cells in ws.columns:
        length = 0
        col_idx = col_cells[0].column
        for cell in col_cells:
            value = cell.value
            if value is None:
                continue
            value_str = str(value)
            length = max(length, len(value_str))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max(length + 2, 12), max_width)


def apply_number_formats(ws, headers: List[str]) -> None:
    header_to_col = {h: i + 1 for i, h in enumerate(headers)}
    currency_headers = {
        "OR Input $/M",
        "OR Output $/M",
        "OR 3:1 Blended $/M",
        "OR $ per AA Point",
        "AA Cheapest Blended $/M",
        "Cheapest Blended $/M",
        "Input $/M Tokens",
        "Output $/M Tokens",
        "Audio Input $/M",
        "Web Search $/K",
        "Lowest Input $/M",
        "Lowest Output $/M",
        "Cost/Test $",
        "Avg Cost Input $",
        "Avg Cost Output $",
    }
    decimal_headers = {
        "AA Preferred Intelligence",
        "AA Alt Intelligence",
        "Intelligence Index",
        "AA Score per Cheapest $",
        "AA Score per OpenRouter $",
        "Fastest Tokens/sec",
        "Lowest Latency (s)",
        "Latency (s)",
        "Accuracy per Cost/Test",
        "AA Points per OR $",
        "Vals Accuracy per $Test",
    }
    pct_headers = {"Vals Accuracy %", "CI +/-", "Percentile from Top"}

    for header, col in header_to_col.items():
        if header in currency_headers:
            fmt = "$0.00"
        elif header in decimal_headers:
            fmt = "0.00"
        elif header in pct_headers:
            fmt = "0.00%"
        elif "Date" in header:
            fmt = "yyyy-mm-dd"
        else:
            continue
        for row in range(2, ws.max_row + 1):
            ws.cell(row, col).number_format = fmt


def build_overview_rows(
    specs: List[ModelSpec],
    pricing_rows: Dict[str, Dict[str, Any]],
    aa_rows: List[Dict[str, Any]],
    vals_rows: List[Dict[str, Any]],
) -> List[Dict[str, Any]]:
    aa_by_family: Dict[str, List[Dict[str, Any]]] = {}
    for row in aa_rows:
        aa_by_family.setdefault(row["Model Family"], []).append(row)

    vals_by_family: Dict[str, List[Dict[str, Any]]] = {}
    for row in vals_rows:
        vals_by_family.setdefault(row["Model Family"], []).append(row)

    rows: List[Dict[str, Any]] = []
    for spec in specs:
        pricing = pricing_rows[spec.family]
        aa_family_rows = aa_by_family.get(spec.family, [])
        vals_family_rows = vals_by_family.get(spec.family, [])

        aa_preferred = next((r for r in aa_family_rows if r.get("_preferred")), aa_family_rows[0] if aa_family_rows else {})
        aa_alt = next((r for r in aa_family_rows if r is not aa_preferred), {})
        vals_preferred = next((r for r in vals_family_rows if r.get("_preferred")), vals_family_rows[0] if vals_family_rows else {})

        rows.append(
            {
                "Model Family": spec.family,
                "OpenRouter Slug": pricing["OpenRouter Slug"],
                "OR Input $/M": pricing["Input $/M Tokens"],
                "OR Output $/M": pricing["Output $/M Tokens"],
                "OR Context": pricing["Context Tokens"],
                "OR Release Date": pricing["Release Date"],
                "AA Preferred Variant": aa_preferred.get("Variant"),
                "AA Preferred Intelligence": aa_preferred.get("Intelligence Index"),
                "AA Alt Variant": aa_alt.get("Variant"),
                "AA Alt Intelligence": aa_alt.get("Intelligence Index"),
                "AA Fastest Provider": aa_preferred.get("Fastest Provider"),
                "AA Fastest t/s": aa_preferred.get("Fastest Tokens/sec"),
                "AA Lowest Latency Provider": aa_preferred.get("Lowest Latency Provider"),
                "AA Lowest Latency (s)": aa_preferred.get("Lowest Latency (s)"),
                "AA Cheapest Provider": aa_preferred.get("Cheapest Provider"),
                "AA Cheapest Blended $/M": aa_preferred.get("Cheapest Blended $/M"),
                "Vals Preferred Variant": vals_preferred.get("Variant"),
                "Vals Accuracy %": vals_preferred.get("Accuracy %"),
                "Vals Latency (s)": vals_preferred.get("Latency (s)"),
                "Vals Cost/Test $": vals_preferred.get("Cost/Test $"),
                "Vals Context": vals_preferred.get("Context Tokens"),
                "Notes": spec.notes,
            }
        )
    return rows


def build_workbook(
    out_path: Path,
    specs: List[ModelSpec],
    pricing_rows: Dict[str, Dict[str, Any]],
    aa_rows: List[Dict[str, Any]],
    vals_rows: List[Dict[str, Any]],
    benchmark_rows: List[Dict[str, Any]],
) -> None:
    wb = Workbook()
    wb.remove(wb.active)

    # Overview
    overview_headers = [
        "Model Family", "OpenRouter Slug", "OR Input $/M", "OR Output $/M", "OR Context",
        "OR Release Date", "OR 3:1 Blended $/M", "AA Preferred Variant", "AA Preferred Intelligence",
        "AA Alt Variant", "AA Alt Intelligence", "AA Fastest Provider", "AA Fastest t/s",
        "AA Lowest Latency Provider", "AA Lowest Latency (s)", "AA Cheapest Provider",
        "AA Cheapest Blended $/M", "Vals Preferred Variant", "Vals Accuracy %", "Vals Latency (s)",
        "Vals Cost/Test $", "Vals Context", "OR $ per AA Point", "AA Points per OR $",
        "Vals Accuracy per $Test", "Notes"
    ]
    overview_rows = build_overview_rows(specs, pricing_rows, aa_rows, vals_rows)
    ws = wb.create_sheet("Overview")
    write_sheet(
        ws,
        overview_headers,
        overview_rows,
        formula_columns={
            "OR 3:1 Blended $/M": "=(C{row}*3+D{row})/4",
            "OR $ per AA Point": '=IFERROR(G{row}/I{row},"")',
            "AA Points per OR $": '=IFERROR(I{row}/G{row},"")',
            "Vals Accuracy per $Test": '=IFERROR(S{row}/U{row},"")',
        },
    )
    apply_number_formats(ws, overview_headers)
    autosize_columns(ws)

    # Pricing_OpenRouter
    pricing_headers = [
        "Model Family", "OpenRouter Slug", "Release Date", "Context Tokens", "Input $/M Tokens",
        "Output $/M Tokens", "Audio Input $/M", "Web Search $/K", "3:1 Blended $/M", "Source URL", "Notes"
    ]
    ws = wb.create_sheet("Pricing_OpenRouter")
    write_sheet(
        ws,
        pricing_headers,
        list(pricing_rows.values()),
        formula_columns={"3:1 Blended $/M": "=(E{row}*3+F{row})/4"},
    )
    apply_number_formats(ws, pricing_headers)
    autosize_columns(ws)

    # ArtificialAnalysis
    aa_headers = [
        "Model Family", "Variant", "Intelligence Index", "Fastest Provider", "Fastest Tokens/sec",
        "Lowest Latency Provider", "Lowest Latency (s)", "Cheapest Provider", "Cheapest Blended $/M",
        "Lowest Input $/M", "Lowest Output $/M", "JSON Support", "Function Calling",
        "Provider Page URL", "Intelligence Source URL", "Notes", "AA Score per Cheapest $", "AA Score per OpenRouter $"
    ]
    aa_export_rows = [{k: v for k, v in row.items() if not k.startswith("_")} for row in aa_rows]
    ws = wb.create_sheet("ArtificialAnalysis")
    write_sheet(
        ws,
        aa_headers,
        aa_export_rows,
        formula_columns={
            "AA Score per Cheapest $": '=IFERROR(C{row}/I{row},"")',
            "AA Score per OpenRouter $": '=IFERROR(C{row}/VLOOKUP(A{row},Pricing_OpenRouter!A:I,9,FALSE),"")',
        },
    )
    apply_number_formats(ws, aa_headers)
    autosize_columns(ws)

    # ValsAI
    vals_headers = [
        "Model Family", "Variant", "Release Date", "Accuracy %", "CI +/-", "Latency (s)", "Cost/Test $",
        "Avg Cost Input $", "Avg Cost Output $", "Context Tokens", "Max Output Tokens",
        "Default Provider", "Reasoning Effort", "Vals Index Rank", "Vals Index Population",
        "Benchmark Highlights", "Source URL", "Accuracy per Cost/Test", "Benchmark Note"
    ]
    vals_export_rows = [{k: v for k, v in row.items() if not k.startswith("_")} for row in vals_rows]
    ws = wb.create_sheet("ValsAI")
    write_sheet(
        ws,
        vals_headers,
        vals_export_rows,
        formula_columns={"Accuracy per Cost/Test": '=IFERROR(D{row}/G{row},"")'},
    )
    apply_number_formats(ws, vals_headers)
    autosize_columns(ws)

    # Vals_Benchmarks
    benchmark_headers = [
        "Model Family", "Variant", "Benchmark", "Rank", "Population", "Percentile from Top", "Source URL", "Notes"
    ]
    ws = wb.create_sheet("Vals_Benchmarks")
    write_sheet(
        ws,
        benchmark_headers,
        benchmark_rows,
        formula_columns={"Percentile from Top": '=IFERROR(1-(D{row}-1)/(E{row}-1),"")'},
    )
    apply_number_formats(ws, benchmark_headers)
    autosize_columns(ws)

    # Sources_Notes
    ws = wb.create_sheet("Sources_Notes")
    notes_rows = [
        ["Local LLM market snapshot", "", "", "", ""],
        ["Generated at (UTC)", now_utc_iso(), "", "", ""],
        ["Workbook builder", "regenerate_model_workbook.py", "", "", ""],
        ["Primary pricing/source universe", "OpenRouter models API + model pages", "https://openrouter.ai/api/v1/models", "https://openrouter.ai/docs/api/api-reference/models/get-models", ""],
        ["Primary benchmark source", "Artificial Analysis free API", "https://artificialanalysis.ai/api/v2/data/llms/models", "https://artificialanalysis.ai/api-reference", ""],
        ["Provider-breakout source", "Artificial Analysis provider pages", "https://artificialanalysis.ai/models/", "", ""],
        ["Application-style metrics source", "Vals AI public model pages", "https://www.vals.ai/models", "https://www.vals.ai/methodology", ""],
        ["Determinism rule", "Cached responses are stored in the cache directory; reruns without --refresh reuse them.", "", "", ""],
        ["Mapping rule", "Cross-source joins are explicit and versioned through the model map.", "", "", ""],
        ["Important caveat", "Release date is scraped from OpenRouter public pages, not taken from the OpenRouter API 'created' field.", "", "", ""],
        ["Important caveat", "Vals page layouts may evolve; review regex/parser if extraction starts missing fields.", "", "", ""],
        ["Important caveat", "Artificial Analysis recommends stable model/creator IDs as primary identifiers.", "", "", ""],
        ["Sheet guide", "Overview", "Curated decision sheet with derived ratios", "", ""],
        ["Sheet guide", "Pricing_OpenRouter", "OpenRouter pricing / context / release snapshot", "", ""],
        ["Sheet guide", "ArtificialAnalysis", "AA model-level + provider-breakout metrics", "", ""],
        ["Sheet guide", "ValsAI", "Vals model-guide metrics", "", ""],
        ["Sheet guide", "Vals_Benchmarks", "Per-benchmark ranks scraped from Vals model pages", "", ""],
    ]
    for row in notes_rows:
        ws.append(row)
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = WHITE_FONT
        cell.alignment = Alignment(horizontal="left", vertical="center")
    autosize_columns(ws, max_width=80)

    # Light styling across sheets.
    for sheet in wb.worksheets:
        sheet.sheet_view.showGridLines = False
        for row in sheet.iter_rows(min_row=2):
            for cell in row:
                if isinstance(cell.value, str) and cell.value.startswith("http"):
                    cell.font = LINK_FONT
                elif isinstance(cell.value, str) and cell.value.startswith("="):
                    cell.font = FORMULA_FONT

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)


# ---------------------------
# CLI / main
# ---------------------------

def main() -> int:
    parser = argparse.ArgumentParser(description="Regenerate the OpenRouter / Artificial Analysis / Vals workbook.")
    parser.add_argument("--out", type=Path, default=Path("openrouter_model_pricing_performance_regenerated.xlsx"))
    parser.add_argument("--cache-dir", type=Path, default=Path(".cache_model_workbook"))
    parser.add_argument("--mapping-csv", type=Path, default=None, help="Optional CSV crosswalk to replace the built-in DEFAULT_MODEL_MAP.")
    parser.add_argument("--refresh", action="store_true", help="Re-fetch remote sources instead of using cached files.")
    args = parser.parse_args()

    specs = load_mapping_csv(args.mapping_csv)
    fetcher = Fetcher(cache_dir=args.cache_dir, refresh=args.refresh)

    pricing_rows = collect_openrouter_data(fetcher, specs)
    aa_rows = collect_aa_data(fetcher, specs)
    vals_rows, benchmark_rows = collect_vals_data(fetcher, specs)

    build_workbook(
        out_path=args.out,
        specs=specs,
        pricing_rows=pricing_rows,
        aa_rows=aa_rows,
        vals_rows=vals_rows,
        benchmark_rows=benchmark_rows,
    )

    print(f"Wrote workbook to: {args.out}")
    print(f"Cached source payloads in: {args.cache_dir}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
