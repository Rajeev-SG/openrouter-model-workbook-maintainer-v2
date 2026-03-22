from pathlib import Path

from openpyxl import load_workbook

from model_intel.workbook.builder import build_workbook


def test_build_workbook_handles_nested_values(tmp_path: Path) -> None:
    out_path = tmp_path / "model-intelligence.xlsx"
    cohort_rows = [
        {
            "canonical_model_id": "deepseek-v3.2-thinking",
            "canonical_family": "DeepSeek V3.2",
            "canonical_variant": "Reasoning / Thinking",
            "provider": "deepseek",
            "reasoning_mode": "reasoning",
            "openrouter_input_price_per_million": 0.26,
            "openrouter_output_price_per_million": 0.38,
            "openrouter_blended_price_per_million": 0.29,
            "openrouter_context_tokens": 163840,
            "aa_release_date": "2026-03-20",
            "aa_intelligence_index": 50.0,
            "aa_coding_index": 36.7,
            "aa_math_index": 42.0,
            "aa_gpqa": 61.0,
            "aa_livecodebench": 57.0,
            "aa_median_tokens_per_second": 41.2,
            "aa_median_ttft_seconds": 0.8,
            "aa_median_ttfat_seconds": 1.4,
            "aa_input_price_per_million": 0.3,
            "aa_output_price_per_million": 0.6,
            "aa_blended_price_per_million": 0.38,
            "aa_fastest_provider": "OpenRouter",
            "aa_fastest_tokens_per_second": 49.0,
            "aa_json_support": "5 / 5",
            "vals_accuracy": 58.0,
            "vals_latency_seconds": 240.0,
            "vals_cost_per_test": 0.14,
            "livebench_overall_score": 71.4,
            "coverage_score": 1.0,
            "has_openrouter": True,
            "has_aa": True,
            "has_vals": True,
            "has_livebench": True,
            "cohort_eligible": True,
            "strict_cohort_eligible": True,
            "exclusion_reasons": "",
            "vals_benchmarks": [{"benchmark": "Vals Index", "rank": 1}],
            "livebench_categories": {"coding": 75.0},
            "livebench_tasks": {"math": 70.0},
            "source_freshness": {"vals_last_seen": "2026-03-17"},
        }
    ]
    scenario_rows = [
        {
            "canonical_model_id": "deepseek-v3.2-thinking",
            "scenario_profile": "coding",
            "scenario_label": "Coding-heavy",
            "scenario_score": 0.91,
            "explanation": {"coding": {"weight": 0.4, "normalized_input": 1.0, "contribution": 0.4}},
        }
    ]
    diagnostics = [{"type": "unmatched-livebench", "source_key": "foo"}]
    source_manifest = {"vals_bundle": {"fetched_at": "2026-03-17T00:00:00+00:00", "record_count": 157}}

    build_workbook(out_path, cohort_rows, cohort_rows, scenario_rows, diagnostics, source_manifest)

    workbook = load_workbook(out_path)

    assert out_path.exists()
    assert "Overview" in workbook.sheetnames
    assert "AA_Benchmarks" in workbook.sheetnames
    assert "AA_Providers" in workbook.sheetnames
    assert "Vals_Benchmarks" in workbook.sheetnames
    assert "LiveBench_Categories" in workbook.sheetnames
