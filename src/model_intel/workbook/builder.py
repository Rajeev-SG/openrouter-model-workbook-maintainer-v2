from __future__ import annotations

import json
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from ..helpers import iso_now


HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
WHITE_FONT = Font(color="FFFFFF", bold=True)
THIN = Side(style="thin", color="D9D9D9")
BOTTOM = Border(bottom=THIN)


def build_workbook(
    out_path: Path,
    cohort_rows: list[dict[str, Any]],
    master_rows: list[dict[str, Any]],
    scenario_rows: list[dict[str, Any]],
    diagnostics: list[dict[str, Any]],
    source_manifest: dict[str, Any],
) -> None:
    workbook = Workbook()
    workbook.remove(workbook.active)

    overview_rows = [
        {
            "Model": row["canonical_family"],
            "Variant": row["canonical_variant"],
            "Provider": row["provider"],
            "Reasoning Mode": row["reasoning_mode"],
            "OpenRouter Input $/M": row.get("openrouter_input_price_per_million"),
            "OpenRouter Output $/M": row.get("openrouter_output_price_per_million"),
            "OpenRouter Blended $/M": row.get("openrouter_blended_price_per_million"),
            "Context Tokens": row.get("openrouter_context_tokens"),
            "AA Intelligence": row.get("aa_intelligence_index"),
            "AA Coding": row.get("aa_coding_index"),
            "AA Tokens/Sec": row.get("aa_median_tokens_per_second"),
            "Vals Accuracy": row.get("vals_accuracy"),
            "Vals Latency (s)": row.get("vals_latency_seconds"),
            "Vals Cost/Test": row.get("vals_cost_per_test"),
            "LiveBench Overall": row.get("livebench_overall_score"),
            "Coverage Score": row.get("coverage_score"),
            "Guide Eligible": row.get("cohort_eligible"),
        }
        for row in cohort_rows
    ]
    coverage_rows = [
        {
            "canonical_model_id": row["canonical_model_id"],
            "canonical_family": row["canonical_family"],
            "provider": row["provider"],
            "reasoning_mode": row["reasoning_mode"],
            "has_openrouter": row["has_openrouter"],
            "has_aa": row["has_aa"],
            "has_vals": row["has_vals"],
            "has_livebench": row["has_livebench"],
            "coverage_score": row["coverage_score"],
            "cohort_eligible": row["cohort_eligible"],
            "exclusion_reasons": row["exclusion_reasons"],
        }
        for row in master_rows
    ]
    exclusion_rows = [row for row in coverage_rows if not row["cohort_eligible"]]
    freshness_rows = [
        {
            "canonical_model_id": row["canonical_model_id"],
            "canonical_family": row["canonical_family"],
            "openrouter_release_date": row.get("openrouter_release_date"),
            "vals_release_date": row.get("vals_release_date"),
            "source_freshness": row.get("source_freshness"),
        }
        for row in master_rows
    ]
    recommendation_rows = _build_recommendation_rows(cohort_rows, scenario_rows)
    vals_benchmark_rows = _flatten_vals_benchmarks(master_rows)
    livebench_category_rows = _flatten_livebench_metrics(master_rows, "livebench_categories", "category", "score")
    livebench_task_rows = _flatten_livebench_metrics(master_rows, "livebench_tasks", "task", "score")

    _write_table(workbook.create_sheet("Overview"), _headers_for(overview_rows), overview_rows)
    _write_table(workbook.create_sheet("Recommendations"), _headers_for(recommendation_rows), recommendation_rows)
    _write_table(workbook.create_sheet("Guide_Cohort"), _headers_for(cohort_rows), cohort_rows)
    _write_table(workbook.create_sheet("Master_Registry"), _headers_for(master_rows), master_rows)
    _write_table(workbook.create_sheet("Coverage"), _headers_for(coverage_rows), coverage_rows)
    _write_table(workbook.create_sheet("Exclusion_Backlog"), _headers_for(exclusion_rows), exclusion_rows)
    _write_table(workbook.create_sheet("Scenario_Scores"), _headers_for(scenario_rows), scenario_rows)
    _write_table(workbook.create_sheet("Vals_Benchmarks"), _headers_for(vals_benchmark_rows), vals_benchmark_rows)
    _write_table(workbook.create_sheet("LiveBench_Categories"), _headers_for(livebench_category_rows), livebench_category_rows)
    _write_table(workbook.create_sheet("LiveBench_Tasks"), _headers_for(livebench_task_rows), livebench_task_rows)
    _write_table(workbook.create_sheet("Source_Freshness"), _headers_for(freshness_rows), freshness_rows)
    _write_table(workbook.create_sheet("Mapping_Audit"), _headers_for(diagnostics), diagnostics)
    _write_sources_sheet(workbook.create_sheet("Sources_Notes"), source_manifest)

    for sheet in workbook.worksheets:
        sheet.sheet_view.showGridLines = False
        _autosize(sheet)

    out_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(out_path)


def _headers_for(rows: list[dict[str, Any]]) -> list[str]:
    if not rows:
        return ["empty"]
    seen: list[str] = []
    for row in rows:
        for key in row.keys():
            if key not in seen:
                seen.append(key)
    return seen


def _build_recommendation_rows(
    cohort_rows: list[dict[str, Any]],
    scenario_rows: list[dict[str, Any]],
) -> list[dict[str, Any]]:
    by_model = {row["canonical_model_id"]: row for row in cohort_rows}
    best_by_profile: dict[str, dict[str, Any]] = {}
    for row in scenario_rows:
        current = best_by_profile.get(row["scenario_profile"])
        if current is None or row["scenario_score"] > current["scenario_score"]:
            best_by_profile[row["scenario_profile"]] = row
    result = []
    for profile, row in sorted(best_by_profile.items()):
        model = by_model.get(row["canonical_model_id"], {})
        result.append(
            {
                "scenario_profile": profile,
                "scenario_label": row.get("scenario_label"),
                "canonical_model_id": row["canonical_model_id"],
                "canonical_family": model.get("canonical_family"),
                "canonical_variant": model.get("canonical_variant"),
                "provider": model.get("provider"),
                "scenario_score": row["scenario_score"],
                "explanation": row.get("explanation"),
            }
        )
    return result


def _flatten_vals_benchmarks(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    flattened = []
    for row in rows:
        for benchmark in row.get("vals_benchmarks") or []:
            flattened.append(
                {
                    "canonical_model_id": row["canonical_model_id"],
                    "canonical_family": row["canonical_family"],
                    "canonical_variant": row["canonical_variant"],
                    "provider": row["provider"],
                    **benchmark,
                }
            )
    return flattened


def _flatten_livebench_metrics(
    rows: list[dict[str, Any]],
    field_name: str,
    label_key: str,
    value_key: str,
) -> list[dict[str, Any]]:
    flattened = []
    for row in rows:
        for label, value in (row.get(field_name) or {}).items():
            flattened.append(
                {
                    "canonical_model_id": row["canonical_model_id"],
                    "canonical_family": row["canonical_family"],
                    "canonical_variant": row["canonical_variant"],
                    label_key: label,
                    value_key: value,
                }
            )
    return flattened


def _write_table(ws, headers: list[str], rows: list[dict[str, Any]]) -> None:
    ws.append(headers)
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = WHITE_FONT
        cell.border = BOTTOM
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for row in rows:
        ws.append([_excel_value(row.get(header)) for header in headers])
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


def _write_sources_sheet(ws, source_manifest: dict[str, Any]) -> None:
    ws.append(["Source", "Fetched At", "Source URL", "Parser Version", "Records", "Failed Count", "Artifact"])
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = WHITE_FONT
        cell.border = BOTTOM
    ws.append(["Generated At", iso_now(), "", "", "", "", ""])
    for name, payload in source_manifest.items():
        ws.append(
            [
                name,
                payload.get("fetched_at"),
                payload.get("source_url"),
                payload.get("parser_version"),
                payload.get("record_count"),
                payload.get("failed_count"),
                payload.get("artifact"),
            ]
        )


def _excel_value(value: Any) -> Any:
    if isinstance(value, (list, dict)):
        return json.dumps(value, ensure_ascii=True, sort_keys=True)
    return value


def _autosize(ws, max_width: int = 42) -> None:
    for column in ws.columns:
        width = max(len(str(cell.value or "")) for cell in column)
        ws.column_dimensions[get_column_letter(column[0].column)].width = min(max(width + 2, 12), max_width)
